#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
导出功能模块
"""

import os
from datetime import datetime
from typing import List, Union, TYPE_CHECKING

if TYPE_CHECKING:
    from src.ui.result_panel import DrawResult

from .database import DrawRecord


class Exporter:
    """导出器"""

    @staticmethod
    def export_to_txt(records: List[DrawRecord], file_path: str) -> bool:
        """导出为 TXT 文件

        Args:
            records: 抽题记录列表
            file_path: 导出文件路径

        Returns:
            是否成功
        """
        try:
            with open(file_path, "w", encoding="utf-8") as f:
                f.write("随机抽题记录\n")
                f.write(f"导出时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"共 {len(records)} 条记录\n")
                f.write("=" * 50 + "\n\n")

                for i, record in enumerate(records, 1):
                    f.write(f"【{i}】{record.question_title}\n")
                    if record.person_name:
                        f.write(f"抽中人员: {record.person_name}\n")
                    f.write(f"题库: {record.bank_name}\n")
                    f.write(f"抽取时间: {record.draw_time.strftime('%Y-%m-%d %H:%M:%S')}\n")
                    if record.question_content:
                        f.write(f"题目要求:\n{record.question_content}\n")
                    f.write("-" * 50 + "\n\n")

            return True
        except Exception:
            return False

    @staticmethod
    def export_to_excel(records: List[DrawRecord], file_path: str) -> bool:
        """导出为 Excel 文件

        Args:
            records: 抽题记录列表
            file_path: 导出文件路径

        Returns:
            是否成功
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side

            wb = Workbook()
            ws = wb.active
            ws.title = "抽题记录"

            # 表头样式
            header_font = Font(bold=True)
            header_alignment = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )

            # 写入表头
            headers = ["序号", "人员", "题目标题", "题库", "抽取时间", "题目要求"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = thin_border

            # 写入数据
            for row, record in enumerate(records, 2):
                ws.cell(row=row, column=1, value=row - 1).border = thin_border
                ws.cell(row=row, column=2, value=record.person_name or "-").border = thin_border
                ws.cell(row=row, column=3, value=record.question_title).border = thin_border
                ws.cell(row=row, column=4, value=record.bank_name).border = thin_border
                ws.cell(row=row, column=5,
                        value=record.draw_time.strftime('%Y-%m-%d %H:%M:%S')).border = thin_border
                ws.cell(row=row, column=6, value=record.question_content).border = thin_border

            # 调整列宽
            ws.column_dimensions["A"].width = 8
            ws.column_dimensions["B"].width = 12
            ws.column_dimensions["C"].width = 35
            ws.column_dimensions["D"].width = 15
            ws.column_dimensions["E"].width = 20
            ws.column_dimensions["F"].width = 45

            wb.save(file_path)
            return True
        except Exception:
            return False

    @staticmethod
    def export_results_to_txt(results: List["DrawResult"], file_path: str) -> bool:
        """导出当前抽题结果为 TXT 文件

        Args:
            results: 抽题结果列表 (DrawResult)
            file_path: 导出文件路径

        Returns:
            是否成功
        """
        try:
            with open(file_path, "w", encoding="utf-8") as f:
                f.write("随机抽题结果\n")
                f.write(f"导出时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"共 {len(results)} 条记录\n")
                f.write("=" * 50 + "\n\n")

                for i, result in enumerate(results, 1):
                    f.write(f"【{i}】{result.question_title}\n")
                    if result.person_name:
                        f.write(f"抽中人员: {result.person_name}\n")
                    f.write(f"题库: {result.bank_name}\n")
                    if result.question_content:
                        f.write(f"题目要求:\n{result.question_content}\n")
                    f.write("-" * 50 + "\n\n")

            return True
        except Exception:
            return False

    @staticmethod
    def export_results_to_excel(results: List["DrawResult"], file_path: str) -> bool:
        """导出当前抽题结果为 Excel 文件

        Args:
            results: 抽题结果列表 (DrawResult)
            file_path: 导出文件路径

        Returns:
            是否成功
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side

            wb = Workbook()
            ws = wb.active
            ws.title = "抽题结果"

            # 表头样式
            header_font = Font(bold=True)
            header_alignment = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )

            # 写入表头
            headers = ["序号", "人员", "题目标题", "题库", "题目要求"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = thin_border

            # 写入数据
            for row, result in enumerate(results, 2):
                ws.cell(row=row, column=1, value=row - 1).border = thin_border
                ws.cell(row=row, column=2, value=result.person_name or "-").border = thin_border
                ws.cell(row=row, column=3, value=result.question_title).border = thin_border
                ws.cell(row=row, column=4, value=result.bank_name).border = thin_border
                ws.cell(row=row, column=5, value=result.question_content).border = thin_border

            # 调整列宽
            ws.column_dimensions["A"].width = 8
            ws.column_dimensions["B"].width = 12
            ws.column_dimensions["C"].width = 35
            ws.column_dimensions["D"].width = 15
            ws.column_dimensions["E"].width = 45

            wb.save(file_path)
            return True
        except Exception:
            return False
